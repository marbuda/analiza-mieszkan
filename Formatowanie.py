import importlib
import utils
import openpyxl
from openpyxl.styles import PatternFill
from utils import *

importlib.reload(utils)


def format_column_color(file_path, col_name, dark_code, light_code, sheet_name=None):
    # Załaduj skoroszyt i wybierz aktywny arkusz
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    # Szukamy kolumny "nazwa"
    nazwa_col = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == col_name:  # Porównujemy z nagłówkiem
            nazwa_col = col[0].column_letter  # Zapamiętujemy literę kolumny
            break

    if nazwa_col is None:
        print("Kolumna 'nazwa' nie została znaleziona.")
        return

    # Ustawienia stylu
    dark_fill = PatternFill(start_color=dark_code, end_color=dark_code, fill_type='solid')  # Ciemnozielony
    light_fill = PatternFill(start_color=light_code, end_color=light_code, fill_type='solid')  # Jasnozielony

    # Formatuj nagłówek
    sheet[f'{nazwa_col}1'].fill = dark_fill

    # Formatuj wszystkie obserwacje w kolumnie "nazwa"
    for row in range(2, sheet.max_row + 1):  # Od drugiego wiersza, aby pominąć nagłówek
        sheet[f'{nazwa_col}{row}'].fill = light_fill

    # Zapisz zmiany w pliku
    workbook.save(file_path)
    print(f"Formatowanie kolumny 'nazwa' zakończone w pliku: {file_path}")


def format_sprzedane_rows(file_path, sprzedane_col='Sprzedane?', sheet_name=None):
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    # Szukamy kolumny "nazwa"
    nazwa_col = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == sprzedane_col:  # Porównujemy z nagłówkiem
            nazwa_col = col[0].column_letter  # Zapamiętujemy literę kolumny
            break
    # Ustal styl wypełnienia
    gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Iteruj przez wiersze w arkuszu
    for row in sheet.iter_rows(min_row=2):  # min_row=2, aby pominąć nagłówki
        cell_value = sheet[f'{nazwa_col}{row[0].row}'].value  # Odwołanie do wartości w kolumnie za pomocą litery
        if cell_value == "TAK":
            for cell in row:
                cell.fill = gray_fill

    # Zapisz zmiany do pliku
    workbook.save(file_path)


def format_duplicates_rows(file_path, duplicate_col = 'Duplicate', sheet_name = None):
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    # Szukamy kolumny "nazwa"
    nazwa_col = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == duplicate_col:  # Porównujemy z nagłówkiem
            nazwa_col = col[0].column_letter  # Zapamiętujemy literę kolumny
            break

    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Iteruj przez wiersze w arkuszu
    for row in sheet.iter_rows(min_row=2):  # min_row=2, aby pominąć nagłówki
        cell_value = sheet[f'{nazwa_col}{row[0].row}'].value  # Odwołanie do wartości w kolumnie za pomocą litery
        if cell_value is not None:
            for cell in row:
                cell.fill = red_fill

    # Zapisz zmiany do pliku
    workbook.save(file_path)


def format_file(path):
    for column, colors in columns_colors_dict.items():
        format_column_color(path, column, colors[0], colors[1])
    format_sprzedane_rows(path)
    format_duplicates_rows(path)
